# -*- coding: utf-8 -*-
import hashlib
from datetime import datetime
from typing import List

import numpy as np
import pandas as pd
import pytz
import streamlit as st

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# =============== è¨­å®š ===============
JST = pytz.timezone("Asia/Tokyo")

# Secrets èª­ã¿è¾¼ã¿ï¼ˆStreamlit Cloud ã® Secrets ã‹ã‚‰ï¼‰
APP_SECRETS = st.secrets.get("app", {})
ALLOWED_DATES: List[str] = list(APP_SECRETS.get("allowed_dates", ["2025-10-25"]))
ALLOWED_PLACES: List[str] = list(APP_SECRETS.get("allowed_places", ["ãƒ¡ã‚¤ãƒ³ã‚¹ãƒ†ãƒ¼ã‚¸"]))
DAY_START = APP_SECRETS.get("day_start", "09:00")
DAY_END = APP_SECRETS.get("day_end", "18:00")
GSHEET_ID = APP_SECRETS.get("gsheet_id", "")
ADMIN_PASSWORD = APP_SECRETS.get("admin_password", "")

# =============== Google Sheets æ¥ç¶š ===============
@st.cache_resource(show_spinner=False)
def get_worksheet():
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"])
    scoped = creds.with_scopes([
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ])
    gc = gspread.authorize(scoped)

    try:
        sh = gc.open_by_key(GSHEET_ID)
    except Exception as ex:
        st.error(f"ã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆã‚’é–‹ã‘ã¾ã›ã‚“ã§ã—ãŸã€‚GSHEET_ID={GSHEET_ID}, Error={ex}")
        raise

    try:
        ws = sh.worksheet("data")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title="data", rows=1000, cols=10)
        ws.append_row(["timestamp", "user_name", "date", "place", "start", "end", "priority"])
    return ws


# =============== ãƒ¦ãƒ¼ãƒ†ã‚£ãƒªãƒ†ã‚£ ===============
def time_slots(day_start: str, day_end: str, step_min: int = 15) -> List[str]:
    base = pd.to_datetime(f"2000-01-01 {day_start}")
    end = pd.to_datetime(f"2000-01-01 {day_end}")
    # ç«¯ç‚¹ã¨ã‚‚ã«å«ã‚€ï¼ˆä¾‹: 09:00, 09:15, ..., 18:00ï¼‰
    return pd.date_range(base, end, freq=f"{step_min}min").strftime("%H:%M").tolist()

SLOTS = time_slots(DAY_START, DAY_END, 15)

def validate_range(start: str, end: str) -> bool:
    return pd.to_datetime(start) < pd.to_datetime(end)

def name_to_color(name: str) -> str:
    """åˆ©ç”¨è€…åã‹ã‚‰å®‰å®šã—ãŸæ·¡è‰² HEX ã‚’ç”Ÿæˆã€‚"""
    palette = [
        "#CFE8FF", "#FFD6A5", "#B9FBC0", "#FFADAD", "#FDFFB6", "#A0C4FF",
        "#CAFFBF", "#9BF6FF", "#F1C0E8", "#BDB2FF", "#FFC6FF", "#E0F7FA",
    ]
    h = int(hashlib.md5(name.encode("utf-8")).hexdigest(), 16)
    return palette[h % len(palette)]

def append_rows(ws, rows: list[list[str]]):
    ws.append_rows(rows, value_input_option="USER_ENTERED")


@st.cache_data(ttl=30)
def load_df() -> pd.DataFrame:
    ws = get_worksheet()  # â† ã“ã“ã§å–å¾—

    records = ws.get_all_records()
    df = pd.DataFrame(records)
    if df.empty:
        df = pd.DataFrame(columns=["timestamp", "user_name", "date", "place", "start", "end", "priority"])
    for c in ["date", "start", "end"]:
        if c in df.columns:
            df[c] = df[c].astype(str)
    if "priority" in df.columns:
        df["priority"] = pd.to_numeric(df["priority"], errors="coerce").astype("Int64")
    return df


def make_excel_by_date(df: pd.DataFrame, date_str: str) -> str:
    """æŒ‡å®šæ—¥ã®ãƒ‡ãƒ¼ã‚¿ã‹ã‚‰ã€å ´æ‰€ã”ã¨ã«ã‚·ãƒ¼ãƒˆã‚’ä½œã‚‹ Excel ã‚’ç”Ÿæˆã—ã€ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ã‚’è¿”ã™ã€‚"""
    df_day = df[df["date"] == date_str].copy()
    if df_day.empty:
        raise ValueError("ã“ã®æ—¥ä»˜ã®ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")

    wb = Workbook()
    wb.remove(wb.active)

    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for place in ALLOWED_PLACES:
        df_p = df_day[df_day["place"] == place].copy()
        ws = wb.create_sheet(title=place[:31])  # ã‚·ãƒ¼ãƒˆåã¯31æ–‡å­—åˆ¶é™

        # ãƒ˜ãƒƒãƒ€ãƒ¼ï¼ˆA1=åˆ©ç”¨è€…, B1ã€œ=æ™‚é–“ã‚¹ãƒ­ãƒƒãƒˆï¼‰
        header = ["åˆ©ç”¨è€…"] + SLOTS
        ws.append(header)
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # å¯¾è±¡åˆ©ç”¨è€…ï¼ˆã“ã®å ´æ‰€ã«å¸Œæœ›ã‚’å‡ºã—ãŸäººï¼‰
        users = df_p["user_name"].dropna().unique().tolist()

        for r, user in enumerate(users, start=2):
            ws.cell(row=r, column=1, value=user)
            ws.cell(row=r, column=1).alignment = Alignment(vertical="center")

            sub = df_p[df_p["user_name"] == user]
            user_color = name_to_color(user)
            fill = PatternFill(start_color=user_color.replace('#',''), end_color=user_color.replace('#',''), fill_type="solid")

            for _, rec in sub.iterrows():
                start, end, pr = str(rec["start"]), str(rec["end"]), int(rec["priority"])
                if not validate_range(start, end):
                    continue
                # é–‹å§‹ãƒ»çµ‚äº†ã®ã‚¹ãƒ­ãƒƒãƒˆ indexï¼ˆçµ‚äº†ã¯â€œå«ã‚ãªã„â€é–‹åŒºé–“ï¼‰
                try:
                    s_idx = SLOTS.index(start)
                    e_idx = SLOTS.index(end)
                except ValueError:
                    # ç¯„å›²å¤–ã¯ã‚¹ã‚­ãƒƒãƒ—
                    continue
                # Excel ã®åˆ—ç•ªå·ï¼ˆA=1, B=2 ...ï¼‰: B åˆ—ãŒ SLOTS[0]
                start_col = 2 + s_idx
                end_col_exclusive = 2 + e_idx  # ã“ã“ã¯å«ã‚ãªã„çµ‚ç«¯

                for c in range(start_col, end_col_exclusive):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = fill
                    label = f"ç¬¬{pr}å¸Œæœ›"
                    cell.value = f"{cell.value},{label}" if cell.value else label
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # ç½«ç·šãƒ»åˆ—å¹…
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_thin
        ws.column_dimensions['A'].width = 18
        for col in range(2, len(SLOTS) + 2):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 4.2

    out_name = f"{date_str.replace('-', '')}.xlsx"
    wb.save(out_name)
    return out_name

# =============== UI ===============
st.set_page_config(page_title="æ–½è¨­åˆ©ç”¨å¸Œæœ›ãƒ•ã‚©ãƒ¼ãƒ ", layout="wide")
st.title("æ–½è¨­åˆ©ç”¨å¸Œæœ› åé›†ãƒ»ç®¡ç†ã‚¢ãƒ—ãƒª")

# Worksheet ã¯ã“ã“ã§ä¸€åº¦å–å¾—ï¼ˆãƒ¦ãƒ¼ã‚¶é€ä¿¡æ™‚ã«ä½¿ç”¨ï¼‰
ws = get_worksheet()

user_tab, admin_tab = st.tabs(["ğŸ“ åˆ©ç”¨è€…ãƒ•ã‚©ãƒ¼ãƒ ", "ğŸ›  ç®¡ç†ï¼ˆä¸€è¦§ãƒ»Excelå‡ºåŠ›ï¼‰"])

with user_tab:
    st.caption("â€» ç¬¬1ã€œç¬¬3å¸Œæœ›ã¯ã™ã¹ã¦å¿…é ˆã§ã™ã€‚æ™‚é–“ã¯15åˆ†åˆ»ã¿ã§é¸æŠã—ã¦ãã ã•ã„ã€‚")

    name = st.text_input("ãŠåå‰ï¼ˆå¿…é ˆï¼‰")

    def hope_block(title: str):
        st.subheader(title)
        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1, 1])
        with c1:
            sel_date = st.selectbox("æ—¥ä»˜", ALLOWED_DATES, key=f"date_{title}")
        with c2:
            sel_place = st.selectbox("å ´æ‰€", ALLOWED_PLACES, key=f"place_{title}")
        with c3:
            start = st.selectbox("é–‹å§‹", SLOTS[:-1], key=f"start_{title}")
        with c4:
            end = st.selectbox("çµ‚äº†", SLOTS[1:], key=f"end_{title}")
        return sel_date, sel_place, start, end

    d1, p1, s1, e1 = hope_block("ç¬¬1å¸Œæœ›")
    d2, p2, s2, e2 = hope_block("ç¬¬2å¸Œæœ›")
    d3, p3, s3, e3 = hope_block("ç¬¬3å¸Œæœ›")

    if st.button("é€ä¿¡ã™ã‚‹", type="primary"):
        errors = []

        name_input = name.strip()

        if not name_input:
            errors.append("ãŠåå‰ã¯å¿…é ˆã§ã™ã€‚")
        else:
            # åå‰ã®é‡è¤‡ãƒã‚§ãƒƒã‚¯ï¼ˆæ­£è¦åŒ–ã—ã¦æ¯”è¼ƒï¼‰
            def normalize_name(s: str) -> str:
                s = str(s).strip().replace("ã€€", " ")
                s = " ".join(s.split())
                return s.lower()

            existing_names = [r.get("user_name", "") for r in ws.get_all_records()]
            existing_norm = {normalize_name(n) for n in existing_names}
            if normalize_name(name_input) in existing_norm:
                errors.append(f"ã“ã®åå‰ã€Œ{name_input}ã€ã¯æ—¢ã«ç™»éŒ²ã•ã‚Œã¦ã„ã¾ã™ã€‚åˆ¥ã®åå‰ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚")

        for idx, (s, e) in enumerate([(s1, e1), (s2, e2), (s3, e3)], start=1):
            if not validate_range(s, e):
                errors.append(f"ç¬¬{idx}å¸Œæœ›ã®æ™‚é–“ç¯„å›²ãŒä¸æ­£ã§ã™ï¼ˆé–‹å§‹ < çµ‚äº†ï¼‰ã€‚")
        if errors:
            st.error("\n".join(errors))
        else:
            ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")
            rows = [
                [ts, name_input, d1, p1, s1, e1, 1],
                [ts, name_input, d2, p2, s2, e2, 2],
                [ts, name_input, d3, p3, s3, e3, 3],
            ]
            try:
                append_rows(ws, rows)
                st.success("é€ä¿¡ã—ã¾ã—ãŸã€‚ã”å”åŠ›ã‚ã‚ŠãŒã¨ã†ã”ã–ã„ã¾ã™ï¼")
                load_df.clear()  # ã‚­ãƒ£ãƒƒã‚·ãƒ¥å‰Šé™¤
            except Exception as ex:
                st.error(f"é€ä¿¡ã«å¤±æ•—ã—ã¾ã—ãŸ: {ex}")

# --- ç®¡ç†ã‚¿ãƒ–ï¼ˆãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ä¿è­·ï¼‰ ---
with admin_tab:
    st.subheader("ç®¡ç†ï¼ˆä¸€è¦§ãƒ»Excelå‡ºåŠ›ï¼‰")

    # ã‚»ãƒƒã‚·ãƒ§ãƒ³ã‚¹ãƒ†ãƒ¼ãƒˆã§èªè¨¼æƒ…å ±ã‚’ç®¡ç†
    if "admin_auth" not in st.session_state:
        st.session_state["admin_auth"] = False
    if "admin_msg" not in st.session_state:
        st.session_state["admin_msg"] = ""

    if st.session_state["admin_auth"]:
        # ãƒ­ã‚°ã‚¢ã‚¦ãƒˆãƒœã‚¿ãƒ³
        col_l, col_r = st.columns([1, 6])
        with col_l:
            if st.button("ãƒ­ã‚°ã‚¢ã‚¦ãƒˆ"):
                st.session_state["admin_auth"] = False
                st.session_state["admin_msg"] = "ãƒ­ã‚°ã‚¢ã‚¦ãƒˆã—ã¾ã—ãŸã€‚"
        with col_r:
            if st.session_state.get("admin_msg"):
                st.info(st.session_state["admin_msg"])

        # èªè¨¼æ¸ˆã¿ãªã‚‰ç®¡ç†ç”»é¢ã‚’è¡¨ç¤º
        df = load_df()
        st.subheader("ãƒ‡ãƒ¼ã‚¿ä¸€è¦§ï¼ˆæœ€æ–°ï¼‰")
        st.dataframe(df, use_container_width=True)

        st.divider()
        st.subheader("Excel å‡ºåŠ›ï¼ˆã‚¬ãƒ³ãƒˆãƒãƒ£ãƒ¼ãƒˆé¢¨ï¼‰")
        selectable_dates = sorted(df["date"].dropna().unique().tolist()) if not df.empty else []
        target_dates = st.multiselect("ä½œæˆã™ã‚‹æ—¥ä»˜ã‚’é¸æŠ", options=selectable_dates, default=selectable_dates)

        if st.button("é¸æŠã—ãŸæ—¥ä»˜ã®Excelã‚’ä½œæˆ"):
            if not target_dates:
                st.warning("å¯¾è±¡æ—¥ä»˜ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
            else:
                for d in target_dates:
                    try:
                        path = make_excel_by_date(df, d)
                        with open(path, "rb") as f:
                            st.download_button(
                                label=f"{d} ã®Excelã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
                                file_name=path,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                data=f.read(),
                            )
                    except Exception as ex:
                        st.error(f"{d} ã®ç”Ÿæˆã«å¤±æ•—: {ex}")

    else:
        # èªè¨¼ãƒ•ã‚©ãƒ¼ãƒ ï¼ˆãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰å…¥åŠ›ï¼‰
        st.info("ç®¡ç†ç”»é¢ã‚’è¡¨ç¤ºã™ã‚‹ã«ã¯ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ãŒå¿…è¦ã§ã™ã€‚")
        pwd = st.text_input("ç®¡ç†ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰", type="password", key="admin_pwd_input")
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("ãƒ­ã‚°ã‚¤ãƒ³"):
                # ç®¡ç†ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ãŒè¨­å®šã•ã‚Œã¦ã„ãªã„å ´åˆã¯è­¦å‘Š
                if not ADMIN_PASSWORD:
                    st.error("ç®¡ç†ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ãŒæœªè¨­å®šã§ã™ã€‚Streamlit ã® Secretsï¼ˆapp.admin_passwordï¼‰ã«ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ã‚’è¨­å®šã—ã¦ãã ã•ã„ã€‚")
                else:
                    if pwd == ADMIN_PASSWORD:
                        st.session_state["admin_auth"] = True
                        st.session_state["admin_msg"] = "èªè¨¼ã«æˆåŠŸã—ã¾ã—ãŸã€‚"
                        st.experimental_rerun()  # èªè¨¼å¾Œã«ç”»é¢ã‚’å†æç”»ã—ã¦ç®¡ç†ç”»é¢ã‚’è¡¨ç¤º
                    else:
                        st.error("ãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰ãŒé•ã„ã¾ã™ã€‚")
        with col2:
            if st.button("ã‚­ãƒ£ãƒ³ã‚»ãƒ«"):
                st.session_state["admin_msg"] = ""
                st.experimental_rerun()
